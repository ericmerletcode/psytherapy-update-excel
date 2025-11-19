from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from io import BytesIO
import re

app = FastAPI()


def parse_export_block(text: str):
    """
    Parse le bloc [EXCEL_EXPORT] et renvoie un dict:
    { nom_section: [ [col1, col2, ...], ... ] }
    """
    sections = {}
    current_section = None
    in_export = False

    for line in text.splitlines():
        line = line.strip()

        if line == "[EXCEL_EXPORT]":
            in_export = True
            continue
        if line == "[/EXCEL_EXPORT]":
            break
        if not in_export:
            continue

        # Nouvelle section [Nom de section]
        m = re.match(r"\[(.+?)\]", line)
        if m:
            current_section = m.group(1)
            sections[current_section] = []
            continue

        # Ligne de données (séparateur |)
        if "|" in line and current_section:
            parts = [c.strip() for c in line.split("|")]
            sections[current_section].append(parts)

    return sections


def append_rows_to_sheet(ws, rows):
    """
    Ajoute des lignes à une feuille openpyxl, en adaptant
    la longueur de la ligne au nombre de colonnes existant.
    """
    max_cols = ws.max_column
    for row in rows:
        if len(row) < max_cols:
            row = row + [""] * (max_cols - len(row))
        elif len(row) > max_cols:
            row = row[:max_cols]
        ws.append(row)


@app.post("/update_excel")
async def update_excel(
    export_block: str = Form(...),
    excel_file: UploadFile = File(...)
):
    # Lecture du fichier Excel en mémoire
    content = await excel_file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Fichier Excel vide ou illisible.")

    try:
        wb = load_workbook(filename=BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Impossible de lire le fichier Excel: {e}")

    # Vérification des onglets obligatoires
    required_sheets = [
        "Infos Patient",
        "Historique Séances",
        "Analyse Clinique",
        "Thérapeutiques",
        "Lithothérapie",
        "Suivi",
    ]
    for name in required_sheets:
        if name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Onglet manquant: {name}")

    ws_infos = wb["Infos Patient"]
    ws_histo = wb["Historique Séances"]
    ws_analyse = wb["Analyse Clinique"]
    ws_therap = wb["Thérapeutiques"]
    ws_litho = wb["Lithothérapie"]
    ws_suivi = wb["Suivi"]

    # Parse du bloc EXCEL_EXPORT
    sections = parse_export_block(export_block)

    # Ajout dans les différents onglets
    if "Historique Séances" in sections and sections["Historique Séances"]:
        append_rows_to_sheet(ws_histo, sections["Historique Séances"])

    if "Analyse Clinique" in sections and sections["Analyse Clinique"]:
        append_rows_to_sheet(ws_analyse, sections["Analyse Clinique"])

    if "Thérapeutiques" in sections and sections["Thérapeutiques"]:
        append_rows_to_sheet(ws_therap, sections["Thérapeutiques"])

    if "Lithothérapie" in sections and sections["Lithothérapie"]:
        append_rows_to_sheet(ws_litho, sections["Lithothérapie"])

    if "Suivi" in sections and sections["Suivi"]:
        append_rows_to_sheet(ws_suivi, sections["Suivi"])

    # Mise à jour Infos Patient : B13–B16
    # sections attendues: [Demande], [Famille], [Sante], [Situation]
    update_cells = {
        "Demande": "B13",
        "Famille": "B14",
        "Sante": "B15",
        "Situation": "B16",
    }

    for section_name, cell in update_cells.items():
        if section_name in sections and sections[section_name]:
            first_row = sections[section_name][0]
            # On suppose que le texte intéressant est dans la dernière colonne de la ligne
            value = first_row[-1].strip()
            ws_infos[cell] = value

    # Sauvegarde en mémoire (en conservant styles et formats)
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="updated.xlsx"'},
    )
